import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_beautiful_excel():
    # 1. Define standard columns
    columns = [
        "Audit Date",
        "Auditor Name",
        "Ticket #",
        "Platform (Android/iOS/Web)",
        "Visual Steps Extraction (Score 1-5)",
        "Environment Accuracy (Score 1-5)",
        "Classification/Routing (Score 1-5)",
        "Video File Attached? (Yes/No)",
        "Phase 1 Latency (s)",
        "Phase 2 Latency (s)",
        "Auditor Remarks / Fix Actions"
    ]
    
    # 2. Pre-populate with sample QA names placeholder rows for auditing
    data = [
        ["2026-05-14", "Manish Sharma", 663524, "Android", 5, 5, 5, "Yes", 4.8, 16.5, "Perfect end-to-end run! Visual steps match video exactly."],
        ["", "QA Evaluator 2", "", "", "", "", "", "", "", "", "Ready for evaluation"],
        ["", "QA Evaluator 3", "", "", "", "", "", "", "", "", "Ready for evaluation"],
        ["", "QA Evaluator 4", "", "", "", "", "", "", "", "", "Ready for evaluation"],
        ["", "QA Evaluator 5", "", "", "", "", "", "", "", "", "Ready for evaluation"],
    ]
    
    # Add an extra 50 empty rows for incoming tickets
    for _ in range(50):
        data.append([""] * len(columns))

    df = pd.DataFrame(data, columns=columns)
    filename = "AI_BugLogger_Quality_Audit_Sheet.xlsx"
    
    # 3. Export to Excel and apply styling
    writer = pd.ExcelWriter(filename, engine="openpyxl")
    df.to_excel(writer, index=False, sheet_name="Audit Tracker")
    
    workbook = writer.book
    worksheet = writer.sheets["Audit Tracker"]
    
    # 4. Styling Configuration
    # Core colors: Midnight Blue for headers, soft blue tint for alternates
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 5. Format headers
    for col_idx, header in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # 6. Auto-fit column widths dynamically
    widths = [15, 20, 12, 20, 18, 18, 18, 15, 15, 15, 50]
    for idx, width in enumerate(widths, start=1):
        col_letter = get_column_letter(idx)
        worksheet.column_dimensions[col_letter].width = width

    # Set header row height higher for readability
    worksheet.row_dimensions[1].height = 40
    
    # 7. Format body rows
    for row_idx in range(2, len(data) + 2):
        worksheet.row_dimensions[row_idx].height = 25
        for col_idx in range(1, len(columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            # Center align numeric columns
            if col_idx in [1, 3, 4, 5, 6, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            # Highlight the first populated row slightly to show sample
            if row_idx == 2:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light Green

    # 8. Freeze panes so headers stay visible
    worksheet.freeze_panes = 'A2'
    
    writer.close()
    print(f"Successfully created premium spreadsheet: {filename}")

if __name__ == "__main__":
    generate_beautiful_excel()
